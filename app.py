import streamlit as st
import pandas as pd
import numpy as np
import re, json, io, hashlib
import plotly.graph_objects as go
from datetime import date, datetime
from dateutil.relativedelta import relativedelta

try:
    from supabase import create_client as _sb_create
    _HAS_SB = True
except ImportError:
    _HAS_SB = False

BUCKET = "ueda-mating"
FILES = {
    "insem": "insemination.csv",
    "mate": "mating_candidates.csv",
    "genomic": "genomic.xlsx",
    "herd": "herd_list.csv",
}

# ══════════════════════════════════════════════════════════════════
#  Supabase storage helpers（カウフローアプリと同一パターンを使用。
#  複数端末（獣医側・農家側）から同じデータを参照・共有するための
#  クラウド共有ストレージとして機能する）
# ══════════════════════════════════════════════════════════════════

@st.cache_resource
def _get_sb():
    if not _HAS_SB:
        return None
    try:
        url = st.secrets["supabase"]["url"]
        key = st.secrets["supabase"]["key"]
        return _sb_create(url, key)
    except Exception as e:
        st.error(f"Supabase接続エラー: {e}")
        return None


def _sb_dl(path):
    sb = _get_sb()
    if not sb:
        return None
    try:
        return sb.storage.from_(BUCKET).download(path)
    except Exception:
        return None


def _sb_ul(path, data, mime="text/csv"):
    sb = _get_sb()
    if not sb:
        st.error("Supabase未接続")
        return False
    try:
        try:
            sb.storage.from_(BUCKET).remove([path])
        except Exception:
            pass
        sb.storage.from_(BUCKET).upload(path, data, {"content_type": mime, "upsert": "true"})
        return True
    except Exception as e:
        st.error(f"保存エラー: {e}")
        return False


def _sb_list():
    sb = _get_sb()
    if not sb:
        return []
    try:
        return sb.storage.from_(BUCKET).list() or []
    except Exception:
        return []


@st.cache_data(ttl=30, show_spinner=False)
def _load_raw_cached(filename, _ver):
    data = _sb_dl(filename)
    if data is None:
        return None, None
    items = {i["name"]: i for i in _sb_list()}
    ts = None
    meta = items.get(filename, {})
    ts_str = meta.get("updated_at", "")
    if ts_str:
        try:
            dt = datetime.fromisoformat(ts_str.replace("Z", "+00:00"))
            ts = dt.strftime("%m/%d %H:%M")
        except Exception:
            pass
    return data, ts


def load_raw(filename):
    return _load_raw_cached(filename, st.session_state.get("upload_ver", 0))


def save_upload(uploaded_file, filename, mime):
    uploaded_file.seek(0)
    return _sb_ul(filename, uploaded_file.read(), mime)


@st.cache_data(ttl=30, show_spinner=False)
def _load_settings_cached(_ver):
    data = _sb_dl("settings.json")
    if data:
        try:
            return json.loads(data.decode("utf-8"))
        except Exception:
            pass
    return {}


def load_settings():
    return _load_settings_cached(st.session_state.get("upload_ver", 0))


def save_settings(d):
    _sb_ul("settings.json", json.dumps(d, ensure_ascii=False).encode("utf-8"), "application/json")


# ══════════════════════════════════════════════════════════════════
#  分類ロジック
# ══════════════════════════════════════════════════════════════════

def classify_semen(remark):
    """授精記録の精液コード(Remark)から品種区分を判定する。"""
    if not isinstance(remark, str):
        return "不明"
    r = remark.strip().upper()
    if not r or r == "-":
        return "不明"
    if r == "ETIVFX":
        return "ホル移植"
    if r.startswith(("HK", "PK", "TW", "AGK")):
        return "F1授精"
    if "ET" in r:
        return "F1移植"
    if r.startswith("WA"):
        return "和牛移植"
    if r[0].isdigit():
        return "ホル雌授精"
    return "不明"


def classify_breed_group(semen_class):
    """品種区分を「ホル・F1・和牛」の3区分にまとめる。"""
    if semen_class in ("ホル雌授精", "ホル移植"):
        return "ホル"
    if semen_class in ("F1授精", "F1移植"):
        return "F1"
    if semen_class == "和牛移植":
        return "和牛"
    return "不明"


def classify_method(b):
    if not isinstance(b, str):
        return "不明"
    b = b.strip().upper()
    return {"N": "自然発情", "E": "移植", "P": "PG授精", "O": "追い移植"}.get(b, "不明")


def classify_method_group(method_class):
    """受胎率分析タブ用の3区分（授精・移植・追い移植）。自然発情とPG授精はいずれもAI（授精）として統合。"""
    if method_class in ("自然発情", "PG授精"):
        return "授精"
    if method_class == "移植":
        return "移植"
    if method_class == "追い移植":
        return "追い移植"
    return "不明"


def classify_result(r):
    if r is None or (isinstance(r, float) and pd.isna(r)):
        return "結果待ち"
    r = str(r).strip().upper()
    if not r or r == "NAN":
        return "結果待ち"
    return {"P": "受胎", "O": "空胎", "A": "流産", "R": "重複授精"}.get(r, "不明")


def parse_date_col(series):
    for fmt in ["%m/%d/%Y", "%m/%d/%y", "%Y/%m/%d", "%Y-%m-%d"]:
        try:
            return pd.to_datetime(series, format=fmt)
        except Exception:
            pass
    return pd.to_datetime(series, format="mixed", errors="coerce")


def normalize_eartag(x):
    """浮動小数点として読み込まれた耳標番号(例: 1.620810e+09)を文字列に正規化する。"""
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return ""
    s = str(x).strip()
    if s.endswith(".0"):
        s = s[:-2]
    try:
        if "." in s or "e" in s.lower():
            s = str(int(round(float(s))))
    except Exception:
        pass
    return s


# ══════════════════════════════════════════════════════════════════
#  データ読み込み・整形
# ══════════════════════════════════════════════════════════════════

def read_csv_auto(raw_bytes):
    for enc in ["utf-8-sig", "cp932", "utf-8"]:
        try:
            return pd.read_csv(io.BytesIO(raw_bytes), encoding=enc)
        except Exception:
            continue
    return pd.read_csv(io.BytesIO(raw_bytes), encoding="utf-8", errors="replace")


def process_insemination(raw_bytes):
    if raw_bytes is None:
        return None
    df = read_csv_auto(raw_bytes)
    df.columns = [c.strip() for c in df.columns]
    rename = {}
    cols = list(df.columns)
    # 元のCSV順序（A〜L相当）に依存しすぎないよう、既知のヘッダ名を優先しつつ位置でも補完
    expect = ["ID", "USDA", "BDAT", "DUE", "Event", "DIM", "Date", "Remark", "R", "T", "B", "Technician"]
    for i, name in enumerate(expect):
        if name not in df.columns and i < len(cols):
            rename[cols[i]] = name
    if rename:
        df = df.rename(columns=rename)

    df["ID"] = df["ID"].astype(str).str.strip()
    df["USDA_norm"] = df["USDA"].apply(normalize_eartag) if "USDA" in df.columns else ""
    df["_bdat"] = parse_date_col(df["BDAT"])
    df["_due"] = parse_date_col(df["DUE"])
    df["_date"] = parse_date_col(df["Date"])
    df["Remark"] = df["Remark"].astype(str).str.strip()
    df["R"] = df["R"].astype(str).str.strip()
    df.loc[df["R"].isin(["nan", "None"]), "R"] = ""
    df["B"] = df["B"].astype(str).str.strip()
    df["Technician"] = df["Technician"].astype(str).str.strip() if "Technician" in df.columns else ""

    df["_semen_class"] = df["Remark"].apply(classify_semen)
    df["_breed_group"] = df["_semen_class"].apply(classify_breed_group)
    df["_method_class"] = df["B"].apply(classify_method)
    df["_method_group"] = df["_method_class"].apply(classify_method_group)
    df["_result_class"] = df["R"].apply(classify_result)
    df["_conc"] = df["_result_class"].map({"受胎": 1, "空胎": 0, "流産": 0})

    # 分析対象フラグ：重複授精(R)は除外
    df["_target"] = df["_result_class"] != "重複授精"
    return df


def process_mating(raw_bytes):
    if raw_bytes is None:
        return None
    for enc in ["cp932", "utf-8-sig", "utf-8"]:
        try:
            df = pd.read_csv(io.BytesIO(raw_bytes), encoding=enc)
            break
        except Exception:
            df = None
    if df is None:
        return None
    df.columns = [c.strip() for c in df.columns]
    cols = list(df.columns)
    # 想定レイアウト: A=雌牛管理番号, C=第1候補, E=第2候補, G=第3候補（B,D,Fは空列）
    out = pd.DataFrame()
    out["eartag"] = df[cols[0]].apply(normalize_eartag)
    for i, label in zip([2, 4, 6], ["候補1", "候補2", "候補3"]):
        if i < len(cols):
            out[label] = df[cols[i]].astype(str).str.strip().replace("nan", "")
        else:
            out[label] = ""
    out = out[out["eartag"] != ""]
    return out.reset_index(drop=True)


# 「値が低いほど良い」形質。この形質は他と逆に、値が低いほどパーセンタイル(pct__)が
# 100%に近づくよう反転して計算する。SCS(体細胞スコア)・SCE(種雄牛難産率)はCDCBの
# 標準的な遺伝評価で「低いほど良い」とされる代表的な形質。
LOWER_IS_BETTER_TRAITS = {"SCS", "SCE"}


def is_lower_better(trait_name):
    return str(trait_name).strip().upper() in LOWER_IS_BETTER_TRAITS


def process_genomic(raw_bytes):
    if raw_bytes is None:
        return None, []
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row = 13
    headers = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]
    headers = [h for h in headers if h is not None]
    n_cols = len(headers)
    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, n_cols + 1)]
        if vals[0] is None:
            continue
        rows.append(vals)
    df = pd.DataFrame(rows, columns=headers)
    id_col = headers[0]
    df = df.rename(columns={id_col: "動物ID"})
    df["動物ID"] = df["動物ID"].astype(str).str.strip()
    df["_id_core"] = df["動物ID"].str.replace("^U", "", regex=True)
    df["_id_kind"] = df["_id_core"].apply(lambda x: "short" if len(x) == 4 else ("long" if len(x) == 10 else "other"))
    trait_cols = [c for c in df.columns if c not in ("動物ID", "Official ID", "レコードの種牡馬NAAB", "_id_core", "_id_kind")]
    for c in trait_cols:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    # 牛群内パーセンタイル順位（1〜100%、通常は値が大きいほど良い形質として計算するが、
    # LOWER_IS_BETTER_TRAITSに該当する形質は値が小さいほど100%に近づくよう反転する）
    for c in trait_cols:
        df[f"pct__{c}"] = df[c].rank(pct=True, ascending=not is_lower_better(c), method="average") * 100
    return df, trait_cols


def process_herd_list(raw_bytes):
    """現在牛群にいる牛の名簿（育成牛リスト等）を読み込む。
    授精記録・ゲノム情報どちらにも登場しない（＝一度も授精されておらず、
    ゲノム情報も長形10桁でしか登録されていない）牛でも、この名簿に牛番号が
    載っていれば検索対象に含められるようにするための補助データ。
    列名は農場によって表記ゆれがありうるため、よくある候補名から自動検出する。
    見つからない場合は1列目を牛番号、2列目を生年月日として扱う。"""
    if raw_bytes is None:
        return None
    df = read_csv_auto(raw_bytes)
    df.columns = [c.strip() for c in df.columns]
    cols = list(df.columns)

    id_candidates = ["ID", "id", "牛番号", "個体番号", "管理番号", "耳標番号"]
    bdat_candidates = ["BDAT", "生年月日", "出生年月日", "BIRTH", "BirthDate"]

    id_col = next((c for c in id_candidates if c in df.columns), cols[0] if cols else None)
    bdat_col = next((c for c in bdat_candidates if c in df.columns), None)
    if bdat_col is None and len(cols) > 1:
        bdat_col = cols[1]

    out = pd.DataFrame()
    out["牛番号"] = df[id_col].apply(normalize_eartag) if id_col else ""
    out["bdat"] = parse_date_col(df[bdat_col]) if bdat_col and bdat_col in df.columns else pd.NaT
    out = out[out["牛番号"] != ""]
    return out.reset_index(drop=True)


def build_cow_master(insem_df, mate_df, genomic_df, herd_df=None):
    """牛番号(farm_id)を軸に、ゲノム情報・交配候補を紐付けたマスタを構築する。
    紐付けロジック（農場の運用ルールに基づく厳密な桁位置指定）：
      1. ゲノムID(短縮形 U+4桁) が牛番号と完全一致
      2. ゲノムID(長形 U+10桁) の 6〜9桁目 が牛番号(4桁)と完全一致
         （例：耳標 1398433135 の6〜9桁目「3313」= 牛番号3313）
      3. 交配候補ファイルの耳標番号(10桁) の 6〜9桁目 が牛番号(4桁)と完全一致
    ※ 以前は「牛番号が耳標番号のどこかに部分文字列として含まれるか」という緩い判定を
      使っていたが、これは桁位置を問わない判定のため、無関係な個体の耳標にたまたま
      同じ数字の並びが含まれるケースで誤紐付けが発生しうる不具合があった。
      桁位置を固定した本ロジックに修正済み。

    検索対象となる牛番号(farm_id)の集合は、①授精記録に登場する牛番号、②ゲノム情報が
    短縮形(4桁)で登録されている牛番号、③（任意）育成牛リスト等の名簿ファイルに
    登場する牛番号、の和集合とする。これにより、まだ一度も授精されていない牛でも、
    ゲノム検査済み、または名簿に登録されていれば検索・表示できる。"""

    def digits_6to9(core):
        """10桁の耳標番号文字列から6〜9桁目(4桁)を取り出す。10桁でなければNoneを返す。"""
        return core[5:9] if len(core) == 10 else None

    genomic_short = {}
    genomic_long_by_farmid = {}
    if genomic_df is not None:
        for _, row in genomic_df.iterrows():
            if row["_id_kind"] == "short":
                genomic_short[row["_id_core"]] = row
            elif row["_id_kind"] == "long":
                key = digits_6to9(row["_id_core"])
                if key:
                    genomic_long_by_farmid[key] = row

    mate_by_farmid = {}
    if mate_df is not None:
        for m in mate_df.itertuples(index=False):
            key = digits_6to9(m.eartag)
            if key:
                mate_by_farmid[key] = m

    herd_bdat_by_farmid = {}
    herd_ids = set()
    if herd_df is not None:
        for h in herd_df.itertuples(index=False):
            herd_ids.add(h.牛番号)
            if pd.notna(h.bdat):
                herd_bdat_by_farmid[h.牛番号] = h.bdat

    insem_ids = set(insem_df["ID"].unique()) if insem_df is not None else set()
    farm_ids = sorted(insem_ids | set(genomic_short.keys()) | herd_ids)
    if not farm_ids:
        return {}, {"genomic_matched": 0, "genomic_unmatched": [], "mate_matched": 0, "mate_unmatched": []}

    master = {}
    diagnostics = {"genomic_matched": 0, "genomic_unmatched": [], "mate_matched": 0, "mate_unmatched": []}
    for fid in farm_ids:
        entry = {"farm_id": fid, "genomic_row": None, "mate_row": None,
                  "herd_bdat": herd_bdat_by_farmid.get(fid)}
        entry["genomic_row"] = genomic_short.get(fid)
        if entry["genomic_row"] is None:
            entry["genomic_row"] = genomic_long_by_farmid.get(fid)
        if entry["genomic_row"] is not None:
            diagnostics["genomic_matched"] += 1
        else:
            diagnostics["genomic_unmatched"].append(fid)

        entry["mate_row"] = mate_by_farmid.get(fid)
        if entry["mate_row"] is not None:
            diagnostics["mate_matched"] += 1
        else:
            diagnostics["mate_unmatched"].append(fid)

        master[fid] = entry
    return master, diagnostics


# ══════════════════════════════════════════════════════════════════
#  ページ設定
# ══════════════════════════════════════════════════════════════════

st.set_page_config(page_title="植田牧場 メイティング・繁殖管理アプリ", layout="wide")
st.markdown("""<style>
.stTabs [data-baseweb="tab-list"]{position:sticky;top:3.2rem;z-index:100;
background-color:white;padding-top:4px;box-shadow:0 2px 4px rgba(0,0,0,.08)}
</style>""", unsafe_allow_html=True)

_S = load_settings()

st.sidebar.title("植田牧場 メイティングアプリ")
st.sidebar.markdown("---")
with st.sidebar.expander("データを更新（CSV取り込み）", expanded=False):
    up_insem = st.file_uploader("① 授精記録（CSV）", type=["csv"], key="up_insem")
    up_mate = st.file_uploader("② 交配精液候補（CSV）", type=["csv"], key="up_mate")
    up_genomic = st.file_uploader("③ ゲノム情報（Excel）", type=["xlsx", "xls"], key="up_genomic")
    up_herd = st.file_uploader("④ 育成牛リスト・名簿（CSV、任意）", type=["csv"], key="up_herd",
                                help="授精記録・ゲノム短縮IDのどちらにも登場しない牛（未授精の育成牛等）"
                                     "も検索対象に含めたい場合にご利用ください。1列目を牛番号、"
                                     "2列目を生年月日として読み込みます。")
    if st.button("保存してデータを共有する", type="primary"):
        saved = []
        if up_insem:
            save_upload(up_insem, FILES["insem"], "text/csv")
            saved.append("授精記録")
        if up_mate:
            save_upload(up_mate, FILES["mate"], "text/csv")
            saved.append("交配精液候補")
        if up_genomic:
            save_upload(up_genomic, FILES["genomic"],
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            saved.append("ゲノム情報")
        if up_herd:
            save_upload(up_herd, FILES["herd"], "text/csv")
            saved.append("育成牛リスト")
        if saved:
            st.session_state["upload_ver"] = st.session_state.get("upload_ver", 0) + 1
            st.cache_data.clear()
            st.success(f"保存完了：{', '.join(saved)}（全端末に共有されます）")
            st.rerun()
        else:
            st.warning("ファイルを選択してください")

# データ読み込み
insem_raw, insem_ts = load_raw(FILES["insem"])
mate_raw, mate_ts = load_raw(FILES["mate"])
genomic_raw, genomic_ts = load_raw(FILES["genomic"])
herd_raw, herd_ts = load_raw(FILES["herd"])

insem_df = process_insemination(insem_raw)
mate_df = process_mating(mate_raw)
genomic_df, trait_cols = process_genomic(genomic_raw)
herd_df = process_herd_list(herd_raw)

cow_master, match_diag = build_cow_master(insem_df, mate_df, genomic_df, herd_df)

st.sidebar.markdown("---")
st.sidebar.markdown("**データ読み込み状況**")
for lbl, ts in [("授精記録", insem_ts), ("交配精液候補", mate_ts), ("ゲノム情報", genomic_ts), ("育成牛リスト", herd_ts)]:
    st.sidebar.markdown(("OK " if ts else "-- ") + lbl + (f"  `{ts}`" if ts else ""))

st.sidebar.markdown("---")
st.sidebar.markdown("**交配精液候補の判定基準**")
st.sidebar.caption("牛番号を検索し直しても、ここで設定した基準は保持されます。")
if trait_cols:
    thresh_trait = st.sidebar.selectbox(
        "基準とする形質", trait_cols,
        index=trait_cols.index("TPI") if "TPI" in trait_cols else 0,
        key="trait_select_mating")
    thresh_val = st.sidebar.slider(
        "牛群内順位の基準（この値以上で候補精液を提示）", 1, 100, 50, 1,
        key="mating_threshold")
else:
    thresh_trait, thresh_val = None, 50
    st.sidebar.info("ゲノム情報を読み込むと設定できます。")

st.sidebar.markdown("---")
st.sidebar.markdown("**表示する形質（追加）**")
st.sidebar.caption("DWP$・NM$・TPI以外に常に表示したい形質を複数選択できます。")
ALWAYS_SHOW_TRAITS = ["DWP$", "NM$", "TPI"]
if trait_cols:
    extra_trait_options = [t for t in trait_cols if t not in ALWAYS_SHOW_TRAITS]
    selected_extra_traits = st.sidebar.multiselect(
        "追加で表示する形質（複数選択可）", extra_trait_options,
        default=["乳量"] if "乳量" in extra_trait_options else [],
        key="trait_multiselect_extra")
else:
    selected_extra_traits = []

st.markdown(
    f"<div style='display:flex;align-items:baseline;gap:12px;flex-wrap:wrap;"
    f"position:sticky;top:0;z-index:200;background:white;padding:8px 0 4px 0;"
    f"box-shadow:0 1px 4px rgba(0,0,0,.07)'>"
    f"<span style='font-size:1.5rem;font-weight:700'>植田牧場 メイティング・繁殖管理アプリ</span>"
    f"</div>", unsafe_allow_html=True)

if insem_df is None:
    st.info("サイドバーから授精記録CSVを保存してください。")
    st.stop()

TABS = st.tabs(["① 牛検索・メイティング", "② 授精成績分析（受胎率）", "データ確認"])

# ══════════════════════════════════════════════════════════════════
#  TAB 1: 牛検索・メイティング
# ══════════════════════════════════════════════════════════════════
with TABS[0]:
    st.subheader("牛検索・メイティング")

    query = st.text_input("牛番号を入力してください", value="", placeholder="例：3620")
    query = query.strip()

    if query:
        if query not in cow_master:
            st.warning(f"牛番号「{query}」の授精記録が見つかりません。番号をご確認ください。")
        else:
            entry = cow_master[query]
            hist = insem_df[insem_df["ID"] == query].sort_values("_date", ascending=False)

            bdat = hist["_bdat"].dropna().iloc[0] if hist["_bdat"].notna().any() else None
            if bdat is None and entry.get("herd_bdat") is not None:
                bdat = entry["herd_bdat"]
            age_months = None
            if bdat is not None:
                delta = relativedelta(date.today(), bdat.date())
                age_months = delta.years * 12 + delta.months

            # ---- 遺伝的順位 ----
            st.markdown("#### 遺伝的順位（牛群内パーセンタイル）")
            grow = entry["genomic_row"]
            if grow is None:
                st.info("この牛のゲノム情報が見つかりません（ゲノム検査未実施、または紐付け未確認の可能性があります）。")
            else:
                gc1, gc2, gc3 = st.columns(3)
                for col, trait in zip([gc1, gc2, gc3], ALWAYS_SHOW_TRAITS):
                    if trait in trait_cols:
                        val = grow.get(trait)
                        pct = grow.get(f"pct__{trait}")
                        label = trait + "（低いほど良い）" if is_lower_better(trait) else trait
                        col.metric(label, f"{val:.0f}" if pd.notna(val) else "-",
                                   f"順位 {pct:.0f}%" if pd.notna(pct) else "")

                if selected_extra_traits:
                    st.markdown("&nbsp;", unsafe_allow_html=True)
                    extra_cols = st.columns(3)
                    for i, trait in enumerate(selected_extra_traits):
                        val = grow.get(trait)
                        pct = grow.get(f"pct__{trait}")
                        label = trait + "（低いほど良い）" if is_lower_better(trait) else trait
                        extra_cols[i % 3].metric(label, f"{val:.2f}" if pd.notna(val) else "-",
                                                  f"順位 {pct:.0f}%" if pd.notna(pct) else "")

            st.markdown("---")

            # ---- 交配精液候補 ----
            # サイドバーで設定した基準（thresh_trait・thresh_val）を用いて、登録済み候補を
            # 表示するか「F1・和牛」と表示するかを判定する。この基準はサイドバーに置かれて
            # いるため、牛番号を検索し直しても設定がリセットされない。
            st.markdown("#### 交配精液候補")
            mrow = entry["mate_row"]
            CAND_STYLE = "font-size:1.35rem;line-height:2.1;margin:4px 0 12px 0"

            if grow is None or not trait_cols or thresh_trait is None:
                st.info("ゲノム情報がないため、順位に基づく判定はできません。")
                if mrow is not None:
                    st.markdown(
                        f"<div style='{CAND_STYLE}'><b>登録済み候補：</b><br>"
                        f"第1候補: <b>{mrow.候補1}</b>　／　第2候補: <b>{mrow.候補2}</b>　／　"
                        f"第3候補: <b>{mrow.候補3}</b></div>", unsafe_allow_html=True)
                else:
                    st.info("この牛の登録済み交配精液候補が見つかりません。")
            else:
                pct = grow.get(f"pct__{thresh_trait}")
                trait_label = thresh_trait + "（低いほど良い）" if is_lower_better(thresh_trait) else thresh_trait
                if pd.isna(pct):
                    st.info(f"{trait_label} のデータがありません。")
                elif pct >= thresh_val:
                    st.success(f"{trait_label} の牛群内順位は {pct:.0f}%（基準 {thresh_val}% 以上）")
                    if mrow is not None:
                        st.markdown(
                            f"<div style='{CAND_STYLE}'>"
                            f"・第1候補: <b>{mrow.候補1}</b><br>"
                            f"・第2候補: <b>{mrow.候補2}</b><br>"
                            f"・第3候補: <b>{mrow.候補3}</b></div>", unsafe_allow_html=True)
                    else:
                        st.warning("登録済み候補精液が見つからないため、別途ご確認ください。")
                else:
                    st.warning(f"{trait_label} の牛群内順位は {pct:.0f}%（基準 {thresh_val}% 未満）→ "
                               f"**F1・和牛**（候補精液の対象外）")

            st.markdown("---")

            # ---- 基本情報 ----
            c1, c2, c3 = st.columns(3)
            c1.metric("牛番号", query)
            c2.metric("月齢", f"{age_months} ヶ月" if age_months is not None else "不明")
            c3.metric("生年月日", bdat.strftime("%Y/%m/%d") if bdat is not None else "不明")

            # ---- 現在の繁殖状況 ----
            st.markdown("#### 現在の繁殖状況")
            target_hist = hist[hist["_target"]].sort_values("_date", ascending=False)
            if len(target_hist) == 0:
                st.info("授精記録がありません（未授精）。")
            else:
                latest = target_hist.iloc[0]
                status = latest["_result_class"]
                if status == "受胎":
                    due = latest["_due"]
                    due_str = due.strftime("%Y/%m/%d") if pd.notna(due) else "不明"
                    st.success(f"受胎中　分娩予定日：{due_str}　"
                               f"（直近授精日 {latest['_date'].strftime('%Y/%m/%d')}／"
                               f"{latest['_semen_class']}／{latest['_method_class']}）")
                elif status == "結果待ち":
                    st.warning(f"結果待ち（授精日 {latest['_date'].strftime('%Y/%m/%d')}／"
                               f"{latest['_semen_class']}／{latest['_method_class']}）")
                else:
                    st.error(f"空胎（直近結果：{status}／"
                             f"授精日 {latest['_date'].strftime('%Y/%m/%d')}／"
                             f"{latest['_semen_class']}／{latest['_method_class']}）")

            with st.expander("授精履歴を表示", expanded=False):
                disp_hist = hist[["_date", "_semen_class", "_method_class", "_result_class", "Technician"]].copy()
                disp_hist.columns = ["授精日", "品種", "方法", "結果", "技術者"]
                disp_hist["授精日"] = disp_hist["授精日"].dt.strftime("%Y/%m/%d")
                st.dataframe(disp_hist.reset_index(drop=True), use_container_width=True)

# ══════════════════════════════════════════════════════════════════
#  TAB 2: 授精成績分析（受胎率）
# ══════════════════════════════════════════════════════════════════
with TABS[1]:
    st.subheader("授精成績分析（精液別受胎率）")

    df = insem_df[insem_df["_target"]].copy()
    df = df[df["_date"].notna()]

    fc1, fc2 = st.columns(2)
    min_d, max_d = df["_date"].min(), df["_date"].max()
    with fc1:
        start_d = st.date_input("授精期間（開始）", value=min_d.date() if pd.notna(min_d) else date(2025, 1, 1))
    with fc2:
        end_d = st.date_input("授精期間（終了）", value=max_d.date() if pd.notna(max_d) else date.today())

    fc3, fc4 = st.columns(2)
    with fc3:
        breed_sel = st.multiselect("品種", ["ホル", "F1", "和牛"], default=["ホル", "F1", "和牛"])
    with fc4:
        method_sel = st.multiselect("授精方法", ["授精", "移植", "追い移植"], default=["授精", "移植", "追い移植"])

    dp = df[(df["_date"].dt.date >= start_d) & (df["_date"].dt.date <= end_d)]
    dp = dp[dp["_breed_group"].isin(breed_sel)]
    dp = dp[dp["_method_group"].isin(method_sel)]

    st.caption(f"集計対象：{len(dp)}件（{start_d} 〜 {end_d}）")

    dp2 = dp[dp["_conc"].notna()]
    if len(dp2) == 0:
        st.info("条件に該当する結果確定済みの記録がありません。")
    else:
        grp = dp2.groupby("Remark").agg(授精頭数=("Remark", "count"), 受胎頭数=("_conc", "sum")).reset_index()
        grp["受胎率"] = grp["受胎頭数"] / grp["授精頭数"] * 100
        grp = grp[grp["授精頭数"] >= 3].sort_values("受胎率", ascending=False)
        if len(grp) == 0:
            st.info("3頭以上の記録がある精液コードがありません。")
        else:
            colors = [("#e74c3c" if r < 30 else "#f39c12" if r < 40 else "#2ecc71") for r in grp["受胎率"]]
            fig = go.Figure(go.Bar(
                x=grp["受胎率"].tolist(), y=grp["Remark"].tolist(), orientation="h",
                marker_color=colors,
                text=(grp["受胎率"].round(1).astype(str) + "%").tolist(), textposition="outside"))
            fig.add_vline(x=30, line_dash="dash", line_color="red", annotation_text="30%",
                          annotation_position="top right")
            fig.add_vline(x=40, line_dash="dash", line_color="orange", annotation_text="40%",
                          annotation_position="top right")
            fig.update_layout(title="精液コード別 受胎率（3頭以上）", height=max(300, len(grp) * 24 + 100),
                              xaxis_title="受胎率(%)", yaxis=dict(autorange="reversed"), margin=dict(r=60))
            st.plotly_chart(fig, use_container_width=True)
            st.dataframe(
                grp.rename(columns={"Remark": "精液コード"}).style.format(
                    {"受胎率": "{:.1f}%", "授精頭数": "{:.0f}", "受胎頭数": "{:.0f}"}),
                use_container_width=True)

# ══════════════════════════════════════════════════════════════════
#  TAB 3: データ確認（紐付け診断）
# ══════════════════════════════════════════════════════════════════
with TABS[2]:
    st.subheader("データ確認・紐付け診断")
    st.markdown("牛番号とゲノム情報・交配候補の自動紐付け結果です。未紐付けの牛は番号の表記ゆれ等をご確認ください。")

    n_total = len(cow_master)
    c1, c2 = st.columns(2)
    c1.metric("ゲノム情報 紐付け済み", f"{match_diag['genomic_matched']} / {n_total} 頭")
    c2.metric("交配候補 紐付け済み", f"{match_diag['mate_matched']} / {n_total} 頭")

    with st.expander("ゲノム情報が未紐付けの牛番号"):
        st.write(match_diag["genomic_unmatched"])
    with st.expander("交配候補が未紐付けの牛番号"):
        st.write(match_diag["mate_unmatched"])

    st.markdown("---")
    st.markdown("#### 元データプレビュー")
    pv1, pv2, pv3, pv4 = st.tabs(["授精記録", "交配精液候補", "ゲノム情報", "育成牛リスト"])
    with pv1:
        if insem_df is not None:
            st.dataframe(insem_df.head(50), use_container_width=True)
    with pv2:
        if mate_df is not None:
            st.dataframe(mate_df.head(50), use_container_width=True)
    with pv3:
        if genomic_df is not None:
            st.dataframe(genomic_df.head(50), use_container_width=True)
    with pv4:
        if herd_df is not None:
            st.dataframe(herd_df.head(50), use_container_width=True)
        else:
            st.info("育成牛リストは未取り込みです（任意ファイルのため必須ではありません）。")
